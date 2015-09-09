from openpyxl.cell import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl import Workbook
from math import ceil
from PIL import Image, ImageOps
import requests
import urllib2
import urllib
import pygame
import string
import time
import PIL
import sys
import os

try:
    os.chdir(os.path.dirname(__file__))
except:
    pass

def bing_search(query, search_type):
    #search_type: Web, Image, News, Video
    key= 'owVBvX3PpdhlSGbRa2Te8MsAPY+gM+ly6DseCkl/Ogk'
    query = urllib.quote(query)
    # create credential for authentication
    user_agent = 'Mozilla/mult.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/mult.0; FDM; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 1.1.mult322)'
    credentials = (':%s' % key).encode('base64')[:-1]
    auth = 'Basic %s' % credentials
    url = 'https://api.datamarket.azure.com/Data.ashx/Bing/Search/'+search_type+'?Query=%27'+query+'%27&$top=5&$format=json'
    headers = {"Authorization" : auth, "User-Agent" : user_agent}
    r = requests.get(url, headers=headers)
    result_list = r.json()['d']['results']
    return result_list

murl = "Web Page Blocked"

idx = 0

keyword = raw_input("Enter a search term: ")

zoom = int(raw_input("Enter zoom from 3 to 1: ")) * 10

if keyword == "temp":
    print "Using cache..."
    ext = ".jpg"
else:
    print "Downloading image..."
    while True:
        murl = bing_search(keyword, 'Image')[idx]["MediaUrl"]
        r = requests.get(murl)
        if not "Web Page Blocked" in r.text:
            break
        print "Blocked. Checking for next image..."
        idx += 1

    ext = murl[murl.rfind("."):]
    print "Caching..."
    with open("temp" + ext, "wb") as temp:
        temp.write(r.content)

    #os.system("temp" + ext)

print "Resizing..."

img = Image.open("temp" + ext).convert('RGB')

maxsize = (int(ceil(7.4*zoom)),int(ceil(3.133*zoom)))

print maxsize

mult = 1

print mult

img.thumbnail(maxsize, PIL.Image.ANTIALIAS)

print "Posterizing..."
#img = ImageOps.posterize(img,4)

maxsize = img.size
img2 = img
start = (0,0)
    
#print "Caching..."
#img.save("temp.png")

#os.system("temp" + ext)

print "Instantiating workbook..."
#wb = load_workbook('art.xltx')
wb = Workbook()
ws = wb.active

print "Initializing Pygame..."
pygame.init()
screen = pygame.display.set_mode(((maxsize[0] * mult) - mult, (maxsize[1] * mult) - mult))
screen.fill((255,255,255))

for i in range(1,maxsize[0]+1):
    for j in range(1,maxsize[1]+1):
        ws.cell(row = i, column = j)

print "Pushing pixels..."

def rgb_to_hex(rgb):
   return '%02x%02x%02x' % rgb

for y in range(start[1],start[1]+img.size[1]):
    for x in range(start[0],start[0]+img.size[0]):
        for event in pygame.event.get():
            if event.type == pygame.QUIT: 
                sys.exit()
        gp = img2.getpixel((x, y))
        if gp[0]==gp[1] and gp[1]==gp[2]:
            if 255-((gp[0]+gp[1]+gp[2])/3) <= 1:
                continue
        screen.fill(pygame.Color(gp[0], gp[1], gp[2], 255), (x*mult,y*mult,mult,mult))
        pygame.display.flip()
        c = ws[get_column_letter(x + 1) + str(y+1)]
        c.value = ""
        hexv = rgb_to_hex((gp[0], gp[1], gp[2]))
        c.fill = PatternFill(start_color=hexv,end_color=hexv,fill_type="solid")


print "Scaling..."
for col in ws.column_dimensions:
    col = ws.column_dimensions[col]
    col.width = 2.76

for row in ws.row_dimensions:
    row = ws.row_dimensions[row]
    row.height = 15



print "Saving..."
wb.save("art.xlsx")
print "Opening..."
os.system("art.xlsx")

sys.exit()
