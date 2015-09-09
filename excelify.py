from flask import Flask, request, redirect, Response
from openpyxl.cell import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from PIL import Image, ImageOps
from openpyxl import Workbook
from math import ceil
import warnings
import socket
import urllib
import PIL
import os

hostaddress = "http://localhost/"

hat = raw_input("Enter the host address (localhost as default, 'ip' for local ip): ")
if hat.startswith("http://"):
    hostaddress = hat
elif hat == "ip":
    hostaddress = "http://" + [ip for ip in socket.gethostbyname_ex(socket.gethostname())[2] if not ip.startswith("127.")][:1][0] + "/"
print "Host address set to " + hostaddress + "."

app = Flask(__name__)

warnings.filterwarnings("ignore")

try:
    os.setcwd(os.path.dirname(__file__))
except:
    pass

def rgb_to_hex(rgb):
   return '%02x%02x%02x' % rgb

@app.route("/<name>.xlsx", methods=["GET"])
def get_excel(name):
    try:
        with open(name + ".xlsx", "rb") as exceldoc:
            resp = Response(exceldoc.read(), status=200, mimetype="application/octet-stream")
            return resp
    except:
        return "sos"



@app.route('/')
def mainpage():
    mainpage = """
<!doctype html>
<title>Excelify!</title>
<h1>Better than an 8-bit Minion</h1>
<form action="excelify" method="post" enctype="multipart/form-data">
File to excelify:
<input type="file" name="image">
<br><br>
Zoom (from 1 to 10):
<input type="number" name="zoom" min="1" max="10" value="5">
<br><br>
<input type="submit" value="Excelify!">
</form>
"""
    resp = Response(mainpage, status=200, mimetype="text/html")
    return resp

@app.route('/excelify', methods=["POST"])
def upload_file():
    try:
        file = request.files["image"]
        if file:
            tn = os.tempnam(os.path.dirname(__file__), "art")
            tn = tn[tn.rfind("\\")+1:] + ".xlsx"
            img = Image.open(file).convert('RGB')
            zoom = 50
            try:
                zoom = int(request.form["zoom"])
                zoom = zoom * 10
            except:
                pass
            maxsize = (int(ceil(7.4*zoom)),int(ceil(3.133*zoom)))
            img.thumbnail(maxsize, PIL.Image.ANTIALIAS)
            maxsize = img.size
            img2 = img
            start = (0,0)
            wb = Workbook()
            ws = wb.active
            for i in range(1,maxsize[0]+1):
                for j in range(1,maxsize[1]+1):
                    ws.cell(row = i, column = j)
            for y in range(start[1],start[1]+img.size[1]):
                for x in range(start[0],start[0]+img.size[0]):
                    gp = img2.getpixel((x, y))
                    if gp[0]==gp[1] and gp[1]==gp[2]:
                        if 255-((gp[0]+gp[1]+gp[2])/3) <= 1:
                            continue
                    c = ws[get_column_letter(x + 1) + str(y+1)]
                    c.value = ""
                    hexv = rgb_to_hex((gp[0], gp[1], gp[2]))
                    c.fill = PatternFill(start_color=hexv,end_color=hexv,fill_type="solid")
            for col in ws.column_dimensions:
                col = ws.column_dimensions[col]
                col.width = 2.76
            for row in ws.row_dimensions:
                row = ws.row_dimensions[row]
                row.height = 15
            c = ws[get_column_letter(start[0]+img.size[0]+1) + str(start[1]+img.size[1]+1)]
            c.value = "Created with Excelify by Milkey Mouse"
            wb.save(tn)
            resp = Response("", status=302)
            resp.headers["Location"] = hostaddress + tn
            return resp
    except:
        return redirect(hostaddress)


if __name__ == '__main__':
    app.run(port=80,host="0.0.0.0")
